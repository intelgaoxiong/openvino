from subprocess import Popen, PIPE
from collections import OrderedDict
from abc import ABC, abstractmethod
from enum import Enum
from dataclasses import dataclass
import time
import ipdb
import re
import xlsxwriter
import collections
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.chart import (AreaChart, Reference, Series)
import sys
import subprocess
import os
import signal
import json

class ExcelEntry:
    pass


class ColumnsMappingBypass(Enum):
    TestNr="TestNr"
    VideoInput="VideoInput"
    FrameDrop="FrameDrop"
    NrOfStreams="NrOfStreams"
    DetFps="DetFps"
    ClassificationFps="ClassificationFps"
    DetNetwork="DetNetwork"
    DetInferReqNr="DetInferReqNr"
    ClassificationNetwork="ClassificationNetwork"
    ClassInferReqNr="ClassInferReqNr"
    ReclassInterval="ReclassInterval"
    Fps="Fps"
    Ips="Ips"

class ColumnsMappingStreaming(Enum):
    TestNr="TestNr"
    VideoInput="VideoInput"  
    NrOfStreams="NrOfStreams"                              
    DetNetwork="DetNetwork"        
    ClassificationNetwork="ClassificationNetwork"                
    Fps="Fps"
    Ips="Ips"    

class ExcelManager:
    
    if str(os.sys.argv[3]).upper() == "BYPASS":
        ColumnsMapping = Enum('ColumnsMapping', {e.name:e.value for e in ColumnsMappingBypass})
    else: 
        ColumnsMapping = Enum('ColumnsMapping', {e.name:e.value for e in ColumnsMappingStreaming})


    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.entries = []      
        self.saved = False  
        self.styles = [{"alignment":Alignment(horizontal='center'), "font":Font(bold=True), "fill":PatternFill("solid", fgColor="D2D1D1"), "border":Border(top=Side(border_style="thin", color="000000"), left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))},
                        {"alignment":Alignment(horizontal='center')}]
        self.BIGGER_THAN_TEXT = 1.35
        self.nextFreeRow = 2;              

    @classmethod
    def getColNumber(cls, arg_colName):        
        return next((idx+1 for idx,data in enumerate(ExcelManager.ColumnsMapping) if data.name == arg_colName), None)
    
    @classmethod
    def getColName(cls, arg_colNumber):
        return next((data.name for idx,data in enumerate(ExcelManager.ColumnsMapping) if (idx+1) == arg_colNumber), None)

    def __writeCell(self, arg_col, arg_row, arg_value, arg_style={}):        
        #Resize the column if necessary
        # TODO: check columns size, one is wrong        
        if int(self.ws.column_dimensions[get_column_letter(arg_col)].width * self.BIGGER_THAN_TEXT) < int(len(str(arg_value)) * self.BIGGER_THAN_TEXT):
            self.ws.column_dimensions[get_column_letter(arg_col)].width = int(len(arg_value) * self.BIGGER_THAN_TEXT)                    
        currentCell = self.ws.cell(column=arg_col, row=arg_row, value=arg_value);        
        for k,v in arg_style.items():
            setattr(currentCell, str(k), v);

    def __createHeader(self, arg_entry:ExcelEntry):        
        # Write in excel every column in arg_entry
        for col, attr in enumerate(arg_entry):                 
            self.__writeCell(col+1, self.nextFreeRow, attr[0], self.styles[0])
        self.nextFreeRow = self.nextFreeRow + 1
        arg_entry.rewind();


    def addEntry(self, arg_entry:ExcelEntry):        
        self.entries.append(arg_entry)


    def __writeDataToExcel(self):        
        self.__createHeader(self.entries[0])        
        for entry in self.entries:
            for colIdx, cell in enumerate(entry):      
                print("WriteCell col row value", colIdx+1, self.nextFreeRow, str(cell[1]))                
                self.__writeCell(colIdx+1, self.nextFreeRow, cell[1], self.styles[1])
            self.nextFreeRow = self.nextFreeRow + 1
        self.__buildCharts();

    def __buildCharts(self):
        # Make some space between charts and table
        # self.nextFreeRow = self.nextFreeRow + 4        
        # print(self.entries[0].getColNumber("DetFps"));
        
        chart = AreaChart()
        chart.title = "Performance"
        chart.style = 13
        chart.x_axis_title = 'Fps'
        chart.y_axis_title = 'Streams'

        
        categories = Reference(self.ws, min_col=ExcelManager.getColNumber(ExcelManager.ColumnsMapping.TestNr.name), min_row=2, max_row=self.nextFreeRow-1)
        data = Reference(self.ws, min_col=ExcelManager.getColNumber(ExcelManager.ColumnsMapping.Fps.name), min_row=2, max_row=self.nextFreeRow-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        self.ws.add_chart(chart, "A"+str(self.nextFreeRow + 4))

        
    def save(self, arg_name):
        if self.saved == False:
            self.__writeDataToExcel()
            self.wb.save(filename=arg_name)
            self.saved == True
        else:
            raise Exception('File has already been saved !')


class Constants:
    DEMO_BASE_PATH = sys.argv[1]
    CONFIGS_BASE_PATH = os.path.realpath(sys.argv[2])    
    OUTPUT_FILE = "./Kpi_results.xlsx"
    TEMP_OUTPUT_FILE = "/tmp/stdoutFile"

    class TermColor:           
        GREEN = '\033[92m'
        YELLOW = '\033[93m'
        RED = '\033[91m'
        BOLD = '\033[1m'
        UNDERLINE = '\033[4m'
        END = '\033[0m'

    

    def pathConcat(arg_basePath, arg_secondaryPath):
        fpath = (arg_basePath[-1] == '/')
        spath = (arg_secondaryPath[0] == '/')

        if fpath:
            if spath:
                return arg_basePath + arg_secondaryPath[1:];
            else:
                return arg_basePath + arg_secondaryPath;
        else:
            if spath:
                return arg_basePath + arg_secondaryPath;
            else:
                return arg_basePath + "/" + arg_secondaryPath;
    

class BypassConfigParser:

    def __init__(self, arg_filePath):
        with open(arg_filePath) as inputConfig:
            self.data = json.load(inputConfig)
        self.mapping = [(ExcelManager.ColumnsMapping.VideoInput.name, self.__getVideoInput),
                        (ExcelManager.ColumnsMapping.FrameDrop.name, self.__getFrameDrop),
                        (ExcelManager.ColumnsMapping.NrOfStreams.name, self.__getNrOfStreams),
                        (ExcelManager.ColumnsMapping.ClassificationNetwork.name, self.__getClassificationNetwork),
                        (ExcelManager.ColumnsMapping.ClassInferReqNr.name, self.__getClassificationInferReq),
                        (ExcelManager.ColumnsMapping.ReclassInterval.name, self.__getReclassificatinInterval),
                        (ExcelManager.ColumnsMapping.DetNetwork.name, self.__getDetectionNetwork),
                        (ExcelManager.ColumnsMapping.DetInferReqNr.name, self.__getDetectionInferReq) ]    
    
    def getData(self):
        return {e[0]:e[1]() for e in self.mapping }
    
    def __getVideoInput(self):        
        videoInput = self.data["scenes"]["video"]["hvaConfig"]["Decode"][0]["Video"]
        return videoInput.split('/')[-1]

    def __getFrameDrop(self):        
        frameDrop = self.data["scenes"]["video"]["hvaConfig"]["FRC"]["DropEveryXFrame"]
        return frameDrop;
    
    def __getNrOfStreams(self):        
        nrOfStreams = self.data["scenes"]["video"]["pipelineNum"];
        return nrOfStreams;
    
    def __getClassificationNetwork(self):        
        network = self.data["scenes"]["video"]["hvaConfig"]["Classification"]["Model"]
        return network.split('/')[-1]
    
    def __getClassificationInferReq(self):
        return self.data["scenes"]["video"]["hvaConfig"]["Classification"]["InferReqNumber"]

    def __getReclassificatinInterval(self):
        return self.data["scenes"]["video"]["hvaConfig"]["Classification"]["ReclassifyInterval"]

    def __getDetectionNetwork(self):
        network = self.data["scenes"]["video"]["hvaConfig"]["Detection"]["Model"]
        return network.split('/')[-1]

    def __getDetectionInferReq(self):
        return self.data["scenes"]["video"]["hvaConfig"]["Detection"]["InferReqNumber"]


class StreamingConfigParser:

    def __init__(self, arg_filePath):
        with open(arg_filePath) as inputConfig:
            self.data = json.load(inputConfig)
        self.mapping = [(ExcelManager.ColumnsMapping.VideoInput.name, self.__getVideoInput),                        
                        (ExcelManager.ColumnsMapping.NrOfStreams.name, self.__getNrOfStreams),
                        (ExcelManager.ColumnsMapping.ClassificationNetwork.name, self.__getClassificationNetwork),
                        (ExcelManager.ColumnsMapping.DetNetwork.name, self.__getDetectionNetwork)]            


    def getData(self):
        return {e[0]:e[1]() for e in self.mapping }
    
    def __getVideoInput(self):        
        videoInput = self.data["scenes"]["pipelineConfig"]["MEDIA_FILE_0"]
        return videoInput.split('/')[-1]
    
    def __getNrOfStreams(self):        
        nrOfStreams = self.data["scenes"]["displayNum"]
        return nrOfStreams;
    
    def __getClassificationNetwork(self):        
        network = self.data["scenes"]["pipelineConfig"]["CLASSIFY_MODEL"]
        return network.split('/')[-1]
        
    def __getDetectionNetwork(self):
        network = self.data["scenes"]["pipelineConfig"]["DETECT_MODEL"]        
        return network.split('/')[-1]


class AppLogParser:
    def __init__(self):                   
        self.extractors = []
        self.result = {}

    def registerExtractors(self, *arg_extractors):
        for ex in arg_extractors:
            self.extractors.append(ex)

    def getResults(self) -> dict:
        for extractor in self.extractors:
            self.result.update(extractor.getResult())
        return self.result;

    def startParsing(self, arg_byteLogFile):        
        fileContent = arg_byteLogFile 
        
        LENGTH = int(len(fileContent));
        START_OFFSET = int(len(fileContent)/1.01);        
        fpsMax = 0        
        # FPS = (0x74, 0x6f, 0x74, 0x61, 0x6c, 0x20, 0x46, 0x50, 0x53, 0x3a);
        # TotalIps = (0x74, 0x6f, 0x74, 0x61, 0x6c, 0x20, 0x49, 0x50, 0x53, 0x3a);
        
        fileContent = fileContent[START_OFFSET:]

        startWhiteSpaces = False
        for b in fileContent:
            if b == 0x20 and startWhiteSpaces == False:
                for extractor in self.extractors:
                    extractor.receiveByte(b)
                startWhiteSpaces = True;
            elif b == 0x20 and startWhiteSpaces == True:
                continue
            elif b != 0x20 :
                startWhiteSpaces = False;
                for extractor in self.extractors:
                    extractor.receiveByte(b)
        

class ExcelEntry:
    
    def __init__(self):
        self.index = 1;
        self.cellValues = {data.name:"None" for data in ExcelManager.ColumnsMapping}


    def __iter__(self):
        return self
    
    def __next__(self):
        if self.index > len(ExcelManager.ColumnsMapping):
            raise StopIteration
        else:           
            print("index = " + str(self.index))
            temp = (ExcelManager.getColName(self.index), self.cellValues[ExcelManager.getColName(self.index)])                                  
            self.index = self.index + 1
            return temp

    def __getitem__(self, k : ExcelManager.ColumnsMapping):
        return self.cellValues[k]
    
    def printMe(self):
        print(self.cellValues)

    def update(self, *args, **kwargs):
        return self.cellValues.update(*args, **kwargs)

    def rewind(self):
        self.index = 1;


class Extractor(ABC):

    @abstractmethod
    def receiveByte(self, arg_byte):
        pass

    @abstractmethod
    def getResult(self) -> dict:
        pass
    

class FpsExtractor(Extractor):

    def __init__(self):
        self.deq = collections.deque([i for i in ('_'*35)])
        self.patternToMatch = "^total\sFPS:\s+(\d+)\.(\d+)"
        self.maxFps = 0
        self.counter = 0;
    
    def receiveByte(self, arg_byte):        
        self.deq.append(chr(arg_byte))
        self.deq.popleft()

        self.counter = self.counter + 1
        if(self.counter == 35):
            self.counter = 0;                                           

        match = re.search(self.patternToMatch, ''.join(self.deq));
        if match:                                                         
            self.maxFps = max(self.maxFps, float(match.group(1)))
            print("MaxFps so far : " + str(self.maxFps))

    def getResult(self):
        return {ExcelManager.ColumnsMapping.Fps.name:self.maxFps}


class IpsExtractor(Extractor):

    def __init__(self):
        self.deq = collections.deque([i for i in ('_'*35)])
        self.patternToMatch = "^total\sIPS:\s+(\d+)\.(\d+)"
        self.maxIps = 0
        self.counter = 0;
    
    def receiveByte(self, arg_byte):        
        self.deq.append(chr(arg_byte))
        self.deq.popleft()

        self.counter = self.counter + 1
        if(self.counter == 35):
            self.counter = 0;                                           

        match = re.search(self.patternToMatch, ''.join(self.deq));
        if match:                                                         
            self.maxIps = max(self.maxIps, float(match.group(1)))
            print("MaxIps so far : " + str(self.maxIps))
    def getResult(self):
        return {ExcelManager.ColumnsMapping.Ips.name:self.maxIps}


class Executor:    
            
    def __removeIfExists(self, arg_filePath):
        if os.path.exists(arg_filePath):
            os.remove(arg_filePath)

    def run(self, arg_script, arg_targetPath, arg_mode):    
        self.__removeIfExists(Constants.TEMP_OUTPUT_FILE)       
        
        with open(Constants.TEMP_OUTPUT_FILE,"wb") as outFile:

            # The os.setsid() is passed in the argument preexec_fn so
            # it's run after the fork() and before  exec() to run the shell.   

            if str(arg_mode).upper() == "BYPASS":
                arg_targetPath = Constants.pathConcat(arg_targetPath, "/bypass/demoUI/")
                proc = subprocess.Popen("cd " + arg_targetPath + "; . ./prepare_demo_no_gui.sh ; ./demo -c "+arg_script , stdout=outFile, shell=True, preexec_fn=os.setsid)
            else:
                arg_targetPath = Constants.pathConcat(arg_targetPath, "/streaming")
                proc = subprocess.Popen("cd " + arg_targetPath + "; . ./env_host.sh; ./kmb_ia_sample/demo -c "+arg_script , stdout=outFile, shell=True, preexec_fn=os.setsid)
                # proc = subprocess.Popen("cd " + arg_targetPath + "; sudo -S ./env_host.sh < " +Constants.PASSWORD_FILE+"; ./kmb_ia_sample/demo -c "+arg_script , stdout=outFile, shell=True, preexec_fn=os.setsid)
                        
            
            print(Constants.TermColor.GREEN + "Demo app will run for 60s ..." + Constants.TermColor.END)
            time.sleep(60);
            print(Constants.TermColor.GREEN + "Stop the demo application" + Constants.TermColor.END)

            # Send the signal to all the process groups
            os.killpg(os.getpgid(proc.pid), signal.SIGINT)
            time.sleep(7)
        with open(Constants.TEMP_OUTPUT_FILE, "rb") as outFile:
            return outFile.read();


def main():        
    executor = Executor()    
    excelManager = ExcelManager()    
    appLogParser = AppLogParser()
    appLogParser.registerExtractors(FpsExtractor(), IpsExtractor())
    
    configFiles = [f for f in os.listdir(Constants.CONFIGS_BASE_PATH) if os.path.isfile(Constants.pathConcat(Constants.CONFIGS_BASE_PATH,f))]

    for idx, confFile in enumerate(configFiles):
        if os.sys.argv[3].upper() == "BYPASS":
            configParser = BypassConfigParser(Constants.pathConcat(Constants.CONFIGS_BASE_PATH, confFile))                
        else:
            configParser = StreamingConfigParser(Constants.pathConcat(Constants.CONFIGS_BASE_PATH, confFile))                        
        outputData = executor.run(Constants.pathConcat(Constants.CONFIGS_BASE_PATH,confFile), Constants.DEMO_BASE_PATH, os.sys.argv[3])
        appLogParser.startParsing(outputData);
        excelEntry = ExcelEntry();
        excelEntry.update(configParser.getData())        
        excelEntry.update(appLogParser.getResults())
        excelEntry.update({ExcelManager.ColumnsMapping.TestNr.name: idx+1})
        if os.sys.argv[3].upper() == "BYPASS":
            excelEntry.update({ExcelManager.ColumnsMapping.DetFps.name: int(excelEntry[ExcelManager.ColumnsMapping.Fps.name]) / int(excelEntry[ExcelManager.ColumnsMapping.FrameDrop.name])})        
            excelEntry.update({ExcelManager.ColumnsMapping.ClassificationFps.name: round(int(excelEntry[ExcelManager.ColumnsMapping.Ips.name]) / int(excelEntry[ExcelManager.ColumnsMapping.DetFps.name]),2)})        
        
        excelManager.addEntry(excelEntry);

    excelManager.save("./TestKpi_report_" + os.sys.argv[3].upper() + ".xlsx")

    
if __name__ == "__main__":
    main()
