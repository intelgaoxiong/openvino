import glob
from inspect import currentframe
import os
import json
from sys import stdout
import re
import subprocess
import enum
import time
import itertools
from threading import Thread
from datetime import datetime
from subprocess import Popen, PIPE
from subprocess import TimeoutExpired
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from enum import auto

increment = (i for i in range(1,128))

class Utilities:    

    TIMEOUT_MESSAGE = "Timeout"    
    EXCEL_OUTPUT_FILE = "BenchmarkResults.xlsx"
    BENCHMARK_CONFIG_FILE = "/tmp/benchmarkConfigFile.json"
    ERROR_STATUS = "FAIL"
    SUCCESS_STATUS = "OK"
    BENCHMARK_TIMEOUT = 480
    NR_OF_NETWORKS_IN_EXCEL = 0
    LOG_DIRECTORY = datetime.now().strftime("%D-%H-%M-%S").replace("/","-")
    FRAMEWORKS = ["onnx", "caffe", "caffe2", "tf", "tf2", "mxnet"]
    FAILING_NETWORKS = [
                "person-vehicle-bike-detection-crossroad-yolov3-1020",
                 "aclnet-des-53-vpu",
                 "vehicle-license-plate-detection-barrier-0106",
                 "vas-face-detection-stage2",
                 "efficientnet-b0",
                 "mobilenet-v3-small-1.0-224",
                 "ssd_mobilenet_v1_fpn_coco",
                 "retinaface-mobilenetv2-0.25-modified",
                 "pedestrian-detection-adas-0002",
                 "face-detection-retail-0004",
                 "face-detection-adas-0001",  ## xing added
                 "regnetx-32gf", ## xing added
                 "ssdlite_mobilenet_edgetpu_coco_quant", ## xing added

                 "efficientnet-b0",
                 "vas-face-detection-stage1",
                 "openpose-pose",
                 "ssd300",
                 "srgan",
                 "action-recognition-0001-encoder",
                 "dla-60xc",
                 "driver-action-recognition-adas-0002-encoder",
                 "face-detection-0200",
                 "face-detection-0202",
                 "face-detection-retail-0005",
                 "facial-landmarks-35-adas-0002",
                 "human-pose-estimation-0001",
                 "landmarks-regression-retail-0009",
                 "person-detection-0202",
                 "person-detection-retail-0013",
                 "person-detection-action-recognition-0005",
                 "person-vehicle-bike-detection-crossroad-1016",
                 "person-detection-raisinghand-recognition-0001",
                 "person-detection-action-recognition-teacher-0002",
                 "semantic-segmentation-adas-0001",
                 "mobilenet-v3-large-1.0-224",
                 "mobilenet-v3-small-1.0-224",
                 "yolo_v5m",
                 "yolo_v5s",
                 "pedestrian-and-vehicle-detector-adas-0001"]
    ALREADY_BENCHMARKED = None

    @classmethod
    def getAlreadyBenchmarked(cls, arg_excelPath):
        alreadyBenchmarked = []
        lastWrittenLine = 0;
        if os.path.exists(arg_excelPath if arg_excelPath.startswith("./") or arg_excelPath.startswith("/") else "./" + arg_excelPath):
            book = openpyxl.load_workbook("./" + arg_excelPath)
            sheet = book.active        
            for i in itertools.count(start=3, step=1):
                currentModel = sheet["A"+str(i)].value
                currentPackage = sheet["B"+str(i)].value
                currentFramework = sheet["C"+str(i)].value
                letter = get_column_letter(Utilities.ExcelHeader.COMPILE_RESULT.value)                
                currentCompResult = sheet[letter+str(i)].value                
                if currentModel is None:
                    break;            
                alreadyBenchmarked.append(tuple(sorted((currentModel,currentPackage,currentFramework))))
                lastWrittenLine = i;
                
        
        return (alreadyBenchmarked, lastWrittenLine)

    @classmethod
    def getInfoAboutNetwork(cls, network_path):        
        lenOfFrameworkFound = 0;

        modelName = ".".join(network_path.split('/')[-1].split(".")[:-1])
        # assert("POR" in network_path or "SCALE" in network_path)
        packageName = "POR" if "POR" in network_path else "SCALE"
        frameworkName = None
        for frmk in Utilities.FRAMEWORKS:
            if frmk in network_path and lenOfFrameworkFound == 0:
                frameworkName = frmk
                lenOfFrameworkFound = len(frmk)
            elif frmk in network_path:                
                if len(frmk) > lenOfFrameworkFound:
                    frameworkName=frmk
                    lenOfFrameworkFound = len(frmk)
            else:
                pass
        assert frameworkName != None, network_path
        return (modelName, packageName, frameworkName)

    @classmethod
    def GetBlobFiles(cls, arg_startingPath):
        blobs = Utilities.pathConcat(arg_startingPath, "/**/FP16-INT8/*.blob")
        blobFiles = glob.glob(blobs, recursive=True) 
        return blobFiles

    @classmethod
    def GetErrorFiles(cls, arg_startingPath):
        errors = Utilities.pathConcat(arg_startingPath, "/**/FP16-INT8/*.err")    
        errorFiles = glob.glob(errors, recursive=True)
        return errorFiles

    @classmethod
    def GetTotalNrOfNetworks(cls, arg_startingPath):
        xmlFiles = Utilities.pathConcat(arg_startingPath, "/**/FP16-INT8/*.xml") 
        return len(glob.glob(xmlFiles, recursive=True))

    @classmethod
    def GetBenchmarkCommand(cls, arg_ovPath:str, arguments:str, device=False):
        cmd = ""
        if device == False:
            cmd = "export LD_LIBRARY_PATH=/opt/intel/vpu_accelerator_samples/hddlunite/lib/:$LD_LIBRARY_PATH; cd " + arg_ovPath + "; source './bin/setupvars.sh' > /dev/null; ./deployment_tools/inference_engine/bin/benchmark_app " + arguments;
        else:
            cmd = "cd " + arg_ovPath + "; source './bin/setupvars.sh' > /dev/null;" + " ./deployment_tools/inference_engine/bin/benchmark_app  -progress" + arguments;

        return cmd;
    

    class ExcelHeader(enum.Enum):
        MODEL = next(increment)
        PACKAGE = next(increment)
        FRAMEWORK = next(increment)
        ISIZE = next(increment)
        COMPILER = next(increment)
        FPS_IREQ1 = next(increment)
        LATENCY_IREQ1 = next(increment)
        FPS_IREQ2 = next(increment)
        LATENCY_IREQ2 = next(increment)
        FPS_IREQ4 = next(increment)
        LATENCY_IREQ4 = next(increment)
        FPS_IREQ8 = next(increment)
        LATENCY_IREQ8 = next(increment)        
        MAX_W = next(increment)        
        TOTAL_W = next(increment)
        IDLE_W = next(increment)
        COMPILE_RESULT = next(increment)
        COMPILE_MESSAGE = next(increment)
        BENCHMARK_MESSAGE = next(increment)    

    @classmethod
    def pathConcat(cls, arg_basePath, arg_secondaryPath):
        fpath = (arg_basePath[-1] == '/')
        spath = (arg_secondaryPath[0] == '/')

        if fpath:
            if spath:
                return arg_basePath + arg_secondaryPath[1:];
            else:
                return arg_basePath + arg_secondaryPath;
        else:
            if spath:
                return arg_basePath + arg_secondaryPath;
            else:
                return arg_basePath + "/" + arg_secondaryPath;
    
    @classmethod
    def insertNewLines(cls, arg_message:str, arg_colWidth:int):
        nrOfInsertedNewLines = 0
        divider = int(len(arg_message) / arg_colWidth) + 1
        offset = int(len(arg_message) / divider)        
        for pos in range(offset, len(arg_message), offset):             
            arg_message = arg_message[:pos] + "\n" + arg_message[pos:]
            nrOfInsertedNewLines += 1
        return (nrOfInsertedNewLines,arg_message)

    @classmethod
    def isPartOfFailingNetworks(cls, arg_blobFile):
        networkName = arg_blobFile.split("/")[-1]
        stopPosition = networkName.rfind("_")        
        networkName = networkName[:stopPosition]
        if networkName in Utilities.FAILING_NETWORKS:
            return True
        return False
    
    @classmethod
    def buildArgumentString(cls,model=None, device="VPUX", nireq=8, config=None):        
        args = (" -m " + model) if model is not None else ""
        args +=" -d " + device
        args +=" -nireq " + str(nireq)
        args += (" --load_config " + config) if config is not None else ""        
        return args;



class PowerReader:
    def __init__(self, device=False):
        self.__running = True;
        self.idle = os.sys.maxsize
        self.max_W = 0
        self.device = device
        self.t = None
        self.INA210_GAIN=200
        self.rails = [
                {
                "name" : "VDDCV",
                "iio_dev_num" : 9,
                "channel" : 0,
                "resistor": 0.0005,
                "voltage": 0.8
                },
                {
                "name" : "DDR_VDDCV",
                "iio_dev_num" : 9,
                "channel" : 1,
                "resistor": 0.003,
                "voltage": 0.8
                },
                # {
                #  "name" : "DDR_VDDQLP",
                #  "iio_dev_num" : 5,
                #  "channel" : 0,
                #  "resistor": 0.01,
                #  "voltage": 0.6
                # },
                {
                "name" : "DDR_VDDQ",
                "iio_dev_num" : 5,
                "channel" : 1,
                "resistor": 0.003,
                "voltage": 1.1
                },
                {
                "name" : "DDR_VAA",
                "iio_dev_num" : 5,
                "channel" : 2,
                "resistor": 0.2,
                "voltage": 1.8
                },
                {
                "name" : "VDDIO_B",
                "iio_dev_num" : 9,
                "channel" : 3,
                "resistor": 0.22,
                "voltage": 3.3
                },
                {
                "name" : "VDDIO",
                "iio_dev_num" : 9,
                "channel" : 2,
                "resistor": 0.003,
                "voltage": 1.8
                },
                {
                "name" : "EMMC_VCORE",
                "iio_dev_num" : 3,
                "channel" : 1,
                "resistor": 0.2,
                "voltage": 0.8
                },
                {
                "name" : "EMMC_VDDQ",
                "iio_dev_num" : 8,
                "channel" : 0,
                "resistor": 0.022,
                "voltage": 1.8
                }
            ]

    def terminate(self):
        self.__running = False;
        print("Waiting for thread to join !")
        self.t.join();
        # TODO close the thread
        print("[INFO] Idle_W : " + str(self.idle))
        return self.max_W, self.idle, self.max_W - self.idle
        return self.max_W - self.idle

    def startReading(self):
        self.t = Thread(target = self.__run)
        self.t.start()     

    def readPowerDevice(self, name, iio_dev_num, channel, resistor, voltage):
        i2c_value = int(subprocess.getoutput("cat /sys/bus/iio/devices/iio:device" + str(iio_dev_num) + "/in_voltage" + str(channel) + "_raw"))
        # print("Raw data read " + str(i2c_value));
        sample_mA = i2c_value / resistor / self.INA210_GAIN
        sample_mW = sample_mA * voltage
        # print("{:10s}: {:4.1f}(RAW) {:4.1f}mA {:4.1f}mW".format(name, i2c_value , sample_mA, sample_mW) )
        return sample_mW           
    
    # #With voltage drop
    # def __run(self):
    #     while self.__running :
    #         counter = 1                               

    #         uV_3p3 = float(subprocess.getoutput("cat /sys/devices/i2c-9/9-004c/hwmon/*/in1_input"))
    #         uV_12p0 = float(subprocess.getoutput("cat /sys/devices/i2c-9/9-004c/hwmon/*/in0_input"))
    #         resistor = 0.005 # [ohms]
    #         rv_12p0  = 12.0 # [V] 
    #         rv_3p3   = 3.3  # [V]

    #         mA_12p0 = uV_12p0 * 1.0 / 1000 / 1000 / resistor
    #         W_12p0  = mA_12p0 * rv_12p0

    #         mA_3p3 = uV_3p3 / 1000 / 1000 / resistor
    #         W_3p3  = mA_3p3 * rv_3p3


    #         self.max_mA = W_3p3 if W_3p3 > self.max_mA else self.max_mA;
    #         self.max_W  = W_12p0 if W_12p0 > self.max_W else self.max_W;
    
    #Reading current
    def __run(self):
        while self.__running :
            try:
                if self.device == False:
                    current_3p3v = float(subprocess.getoutput("cat /sys/devices/i2c-9/9-004c/hwmon/*/curr1_input"))                                    
                    rv_3p3 = float(subprocess.getoutput("cat /sys/devices/i2c-9/9-004c/hwmon/*/in1_input"))            
                    W_3p3 = current_3p3v / 1000 * (rv_3p3 / 1000)            
                    self.idle = min(self.idle, W_3p3)
                    self.max_W = max(self.max_W, W_3p3)
                else:                    
                    # [self.readPowerDevice(**rail) for rail in self.rails]
                    totalW = 0
                    for rail in self.rails:
                        totalW += self.readPowerDevice(**rail)
                    totalW = totalW / 1000
                    self.idle = min(self.idle, totalW)  
                    self.max_W = max(self.max_W, totalW)  

            except ValueError:                
                print("[WARNING] Can't read the power consumption ")                
                self.idle = 0;
                self.max_W = 0;
                        
            time.sleep(1)
    
class ParamsExtractor:

    def __init__(self, arg_jsonFilePath):
        with open(arg_jsonFilePath,"r") as conf:
            self.jsonObject = json.load(conf)

    # Return example: [("d","VPUX"), ("nireq","8")]
    def getParams(self) -> list :
        for element in self.jsonObject.items():  
            if element[0] == "params":                                
                return [(pair[0],pair[1]) for pair in element[1].items()]
    
    def getNireq(self) -> list :
        for element in self.jsonObject.items():  
            if element[0] == "params":                
                temp = [pair[1] for pair in element[1].items() if pair[0] == "nireq"][0]
                result = [int(e) for e in temp.split(",")]
                return result;   

    def getNiter(self) -> int :
        for element in self.jsonObject.items():  
            if element[0] == "params":
                for pair in element[1].items():
                    if pair[0] == "niter":
                        return int(pair[1])

    def getOVPath(self) -> str :
        for element in self.jsonObject.items():  
            if element[0] == "paths":                
                return element[1]["ov_path"]
    
    def getBlobFilesPath(self) -> str :
        for element in self.jsonObject.items():  
            if element[0] == "paths":
                return element[1]["blob_files_path"]
    
    def getConfigFilePath(self) -> str:
        path = None
        for element in self.jsonObject.items():            
            if element[0] == "configFile" and len(element[1]) > 0:
                with open(Utilities.BENCHMARK_CONFIG_FILE, 'w') as benchmarkConfig:
                    json.dump(element[1], benchmarkConfig)
                    path = Utilities.BENCHMARK_CONFIG_FILE
                break;
        return path

    def getHardware(self) -> str:
        for element in self.jsonObject.items():            
            if element[0] == "hardware" and len(element[1]) > 0:
                return element[1].upper()                



# Blob files generator
def blobFilesProvider(blobFilesPath, nextGen):
    # blobs = Utilities.pathConcat(blobFilesPath, "/**/FP16-INT8/*.blob")
    # errors = Utilities.pathConcat(blobFilesPath, "/**/FP16-INT8/*.err")    
    # totalNrOfNetworks = Utilities.pathConcat(blobFilesPath, "/**/FP16-INT8/*.xml")    
    blobFiles = Utilities.GetBlobFiles(blobFilesPath)
    errorFiles = Utilities.GetErrorFiles(blobFilesPath)
    totalNrOfNetworks = Utilities.GetTotalNrOfNetworks(blobFilesPath)
    
    # TODO : Don't remove this line
    #assert((len(blobFiles) + len(errorFiles)) == totalNrOfNetworks)    
         
    for file in itertools.chain(blobFiles,errorFiles):                             
        currentNetwork = tuple(sorted(Utilities.getInfoAboutNetwork(file)))        
        if currentNetwork not in Utilities.ALREADY_BENCHMARKED:    
            nextGen.send(file)
        else:            
            pass
         
    nextGen.close()

# TODO take into consideration params in json file 
def runBenchmark(arg_openVinoPath, paramsExtractor : ParamsExtractor, nextGen, device=False):
    arg_ovPath = arg_openVinoPath
    print("RunBenchmark is initialized ! ")    
    dimRegex = re.compile(".*dimensions\s*\(.*\):\s*([\d|\s]*)")
    latencyRegex = re.compile(".*Latency:\s*([\d|\.]*)\sms")
    throughputRegex = re.compile(".*Throughput:\s*([\d|\.]*)\sFPS")
    nireqMapping = {
                    1:(Utilities.ExcelHeader.FPS_IREQ1,Utilities.ExcelHeader.LATENCY_IREQ1),
                    2:(Utilities.ExcelHeader.FPS_IREQ2,Utilities.ExcelHeader.LATENCY_IREQ2),
                    4:(Utilities.ExcelHeader.FPS_IREQ4,Utilities.ExcelHeader.LATENCY_IREQ4),
                    8:(Utilities.ExcelHeader.FPS_IREQ8,Utilities.ExcelHeader.LATENCY_IREQ8)
                    }

    while True:        
        excelEntry = {e:None for e in Utilities.ExcelHeader}
        # Receive new blob file      
        arg_blobFilePath = yield     
        modelName, packageName, frameworkName = Utilities.getInfoAboutNetwork(arg_blobFilePath)
                
        # if arg_blobFilePath.split("/")[-1] in ["mobilenet-v2.blob","mobilenet-v2_MCM.blob","resnet-50-pytorch_MCM.blob", "ssd_mobilenet_v1_coco_MCM.blob", "unet-camvid-onnx-0001_MCM.blob", "yolo_v4_MCM.blob" ] :         
        if True:            
            # modelName = ".".join([part for part in arg_blobFilePath.split('/')[-1].split(".")[:-1]])   
            compilerType = "MCM" if "MCM" in modelName else "MLIR"            
            localMax_W = 0    
            localTotal_W = 0;   
            localIdle_W = 0;     
            #Update excelEntry with model name and compiler type (MLIR or MCM)
            #Update with framework and package extracted from file name
            excelEntry.update({Utilities.ExcelHeader.MODEL: modelName, Utilities.ExcelHeader.COMPILER: compilerType})
            excelEntry.update({
                Utilities.ExcelHeader.FRAMEWORK: frameworkName, 
                Utilities.ExcelHeader.PACKAGE: packageName
                })            
            
            
            if arg_blobFilePath.endswith(".blob") : 
                if Utilities.isPartOfFailingNetworks(arg_blobFilePath):
                    print("[INFO] Skip "+arg_blobFilePath);
                    excelEntry.update({Utilities.ExcelHeader.COMPILE_RESULT:"Blacklisted", Utilities.ExcelHeader.COMPILE_MESSAGE:"Blacklisted"})    
                    excelEntry.update({Utilities.ExcelHeader.BENCHMARK_MESSAGE:"Blacklisted"})    
                else:
                    excelEntry.update({Utilities.ExcelHeader.COMPILE_RESULT:Utilities.SUCCESS_STATUS, Utilities.ExcelHeader.COMPILE_MESSAGE:"Compilation successful"})
                    print("Run benchmark for " + str(arg_blobFilePath))                    
                    for nireq in paramsExtractor.getNireq():
                        try:                                                                
                            process = subprocess.Popen('/bin/bash', stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)     
                            # arguments =  " -m "+ arg_blobFilePath + " -d VPUX -nireq " + str(nireq)          
                            arguments = Utilities.buildArgumentString(model=arg_blobFilePath, config=paramsExtractor.getConfigFilePath())                            
                            # cmd = "export LD_LIBRARY_PATH=/opt/intel/vpu_accelerator_samples/hddlunite/lib/:$LD_LIBRARY_PATH; cd " + arg_ovPath + "; source './bin/setupvars.sh' > /dev/null; ./deployment_tools/inference_engine/bin/benchmark_app " + arguments;
                            cmd = Utilities.GetBenchmarkCommand(arg_ovPath, arguments, device=device)   
                            print("xing test benchmark_app cmd:",cmd)                         
                            pr = PowerReader(device=device)
                            pr.startReading() 
                            time.sleep(1)                            
                            output, err = process.communicate(cmd.encode(), timeout=Utilities.BENCHMARK_TIMEOUT);
                            time.sleep(5)
                        except TimeoutExpired:                                  
                            process.kill()
                            excelEntry.update({Utilities.ExcelHeader.BENCHMARK_MESSAGE:"Timeout"})
                            print("[ERROR] Timeout in benchmark ")     
                            continue                   
                        
                        tempStdout = arg_blobFilePath.split("/")[-1].split(".")[0]+"_"+str(nireq)+"_stdout"
                        tempStderr = arg_blobFilePath.split("/")[-1].split(".")[0]+"_"+str(nireq)+"_stderr"
                        
                        
                        
                        with open("./"+Utilities.LOG_DIRECTORY+"/"+tempStdout, 'w+') as stdOutFile:
                            stdOutFile.write(output.decode())
                        with open("./"+Utilities.LOG_DIRECTORY+"/"+tempStderr, 'w+') as stdErrFile:
                            stdErrFile.write(err.decode())

                        if "Throughput" in output.decode() and (len(err) < 1 or "WARN" in err.decode()):
                            # Start building the excel entry 
                            message = output.decode().replace('\n',"")
                            dim = dimRegex.match(message).group(1)
                            latency = latencyRegex.match(message).group(1)
                            throughput = throughputRegex.match(message).group(1)
                            excelEntry.update({nireqMapping[nireq][0]:throughput, nireqMapping[nireq][1]:latency})
                            excelEntry.update({Utilities.ExcelHeader.ISIZE:dim})                            
                            excelEntry.update({Utilities.ExcelHeader.BENCHMARK_MESSAGE:Utilities.SUCCESS_STATUS})
                            total_W, idle_W, temp_W = pr.terminate();   
                            print("[INFO] Total_W = " + str(total_W) + " Idle_W = " + str(idle_W))                     
                            localMax_W = max(temp_W, localMax_W)   
                            localTotal_W = max(total_W, localTotal_W)     
                            localIdle_W = idle_W                     
                        elif len(err) > 1 and (("reboot" in output.decode()) or ("NETWORK_NOT_LOADED" in err.decode())):
                            print("[INFO] Exit now !")
                            *_, temp_W = pr.terminate()
                            os.sys.exit(1)
                        else:                    
                            message = err.decode().replace('\n',"")
                            excelEntry.update({Utilities.ExcelHeader.BENCHMARK_MESSAGE:message})
            elif arg_blobFilePath.endswith(".err"):
                with open(arg_blobFilePath, "r") as errFile:
                    excelEntry.update({Utilities.ExcelHeader.COMPILE_RESULT:Utilities.ERROR_STATUS, Utilities.ExcelHeader.COMPILE_MESSAGE:errFile.read()})            
                
            else:
                raise Exception("[ERROR] Unknown input file type in benchmark application !")

            excelEntry.update({Utilities.ExcelHeader.MAX_W: localMax_W})
            excelEntry.update({Utilities.ExcelHeader.TOTAL_W: localTotal_W})
            excelEntry.update({Utilities.ExcelHeader.IDLE_W: localIdle_W})
            nextGen.send(excelEntry)

def excelWriterGen(arg_fileName : str, arg_startLine : int):
    excelFileExists = os.path.exists(arg_fileName)
    if excelFileExists == True :
        wb = openpyxl.load_workbook(arg_fileName)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    entries = []      
    saved = False  
    styles = [{"alignment":Alignment(horizontal='center'), "font":Font(bold=True), "fill":PatternFill("solid", fgColor="D2D1D1"), "border":Border(top=Side(border_style="thin", color="000000"), left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))},
              {"alignment":Alignment(horizontal='center')},
              {"fill":PatternFill("solid", fgColor="D2D1D1"), "font.color.index":"FFFF0000"}]        
    BIGGER_THAN_TEXT = 1.2
    ERROR_MESSAGE_WIDTH = 125
    HEIGHT_OF_ROW = 21
    nextFreeRow = 2    

    # header = ["BlobName", "Result", "Message"];
    header = {
                Utilities.ExcelHeader.MODEL : "ModelName",
                Utilities.ExcelHeader.PACKAGE : "Package",
                Utilities.ExcelHeader.FRAMEWORK : "FW",
                Utilities.ExcelHeader.ISIZE : "InputSize",
                Utilities.ExcelHeader.COMPILER : "Compiler",
                Utilities.ExcelHeader.FPS_IREQ1 : "FPS_NIREQ1",
                Utilities.ExcelHeader.LATENCY_IREQ1 : "Latency_NIREQ1",
                Utilities.ExcelHeader.FPS_IREQ2 : "FPS_NIREQ2",
                Utilities.ExcelHeader.LATENCY_IREQ2 : "Latency_NIREQ2",
                Utilities.ExcelHeader.FPS_IREQ4 : "FPS_NIREQ4",
                Utilities.ExcelHeader.LATENCY_IREQ4 : "Latency_NIREQ4",
                Utilities.ExcelHeader.FPS_IREQ8 : "FPS_NIREQ8",
                Utilities.ExcelHeader.LATENCY_IREQ8 : "Latency_NIREQ8",                
                Utilities.ExcelHeader.MAX_W : "Max_W",
                Utilities.ExcelHeader.TOTAL_W : "Total_W",
                Utilities.ExcelHeader.IDLE_W : "Idle_W",
                Utilities.ExcelHeader.COMPILE_RESULT : "CompResult",
                Utilities.ExcelHeader.COMPILE_MESSAGE: "CompMessage",
                Utilities.ExcelHeader.BENCHMARK_MESSAGE: "BenchmarkMessage"
             }
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.MODEL.value)].width = 50
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.PACKAGE.value)].width = 15
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.FRAMEWORK.value)].width = 15
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.ISIZE.value)].width = 25
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.COMPILER.value)].width = 15
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.FPS_IREQ1.value)].width = 17
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.LATENCY_IREQ1.value)].width = 20
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.FPS_IREQ2.value)].width = 17
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.LATENCY_IREQ2.value)].width = 20
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.FPS_IREQ4.value)].width = 17
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.LATENCY_IREQ4.value)].width = 20
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.FPS_IREQ8.value)].width = 17
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.LATENCY_IREQ8.value)].width = 20    
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.MAX_W.value)].width = 20    
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.TOTAL_W.value)].width = 20    
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.IDLE_W.value)].width = 20    
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.COMPILE_RESULT.value)].width = 15
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.COMPILE_MESSAGE.value)].width = ERROR_MESSAGE_WIDTH    
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.BENCHMARK_MESSAGE.value)].width = ERROR_MESSAGE_WIDTH / 2    

    if excelFileExists == False:
        # Create excel header    
        for idx, e in enumerate(Utilities.ExcelHeader):
            # if int(ws.column_dimensions[get_column_letter(idx+1)].width * BIGGER_THAN_TEXT) < int(len(str(header[idx])) * BIGGER_THAN_TEXT):
            #     ws.column_dimensions[get_column_letter(idx+1)].width = int(len(header[idx]) * BIGGER_THAN_TEXT)        
            cell = ws.cell(column=idx+1, row=nextFreeRow, value=header[e]);
            for k, v in styles[0].items():
                setattr(cell, str(k), v);
        nextFreeRow = nextFreeRow + 1
    else:
        nextFreeRow = arg_startLine
    
    try:
        while True:
            entry = yield             
            printInRed = False;
            Utilities.NR_OF_NETWORKS_IN_EXCEL += 1

            # entry = {} like header
            #Strip all new lines in message
            entry[Utilities.ExcelHeader.COMPILE_MESSAGE] = entry[Utilities.ExcelHeader.COMPILE_MESSAGE].replace("\n","")
            nrOfNewLines, entry[Utilities.ExcelHeader.COMPILE_MESSAGE] = Utilities.insertNewLines(entry[Utilities.ExcelHeader.COMPILE_MESSAGE], ERROR_MESSAGE_WIDTH)            
            ws.row_dimensions[nextFreeRow].height = HEIGHT_OF_ROW *nrOfNewLines

            for idx, value in enumerate(Utilities.ExcelHeader):
                cell = ws.cell(column=idx+1, row=nextFreeRow, value=entry[value]);
                cell.alignment = Alignment(vertical='center', horizontal='center')
                if value == Utilities.ExcelHeader.COMPILE_RESULT and entry[value] != Utilities.SUCCESS_STATUS:  
                    printInRed = True
                    cell.font = Font(color="FFFF0000")     
                if value == Utilities.ExcelHeader.BENCHMARK_MESSAGE and entry[value] != Utilities.SUCCESS_STATUS:
                    printIntRed = True
                    cell.font = Font(color="FFFF0000")                    
            if printInRed:
                for idx, value in enumerate(Utilities.ExcelHeader):
                    cell = ws.cell(column=idx+1, row=nextFreeRow).font = Font(color="FFFF0000");

            nextFreeRow = nextFreeRow + 1  
            wb.save(filename = arg_fileName)          
    finally:
        wb.save(filename = arg_fileName)    

    


def main():
    configFilePath = os.path.realpath(os.sys.argv[1])
    p = ParamsExtractor(configFilePath)
    runOnDevice = True if p.getHardware() == "DEVICE" else False    
    os.mkdir("./"+Utilities.LOG_DIRECTORY)
    Utilities.ALREADY_BENCHMARKED, lastLine = Utilities.getAlreadyBenchmarked(Utilities.EXCEL_OUTPUT_FILE)    

    excelWriter = excelWriterGen(Utilities.EXCEL_OUTPUT_FILE, lastLine + 1); next(excelWriter);
    if os.path.exists(p.getOVPath()) == False: 
        print("[ERROR] OpenVINO path not found !")
    if os.path.exists(p.getBlobFilesPath()) == False: 
        print("[ERROR] Blob files path not found !")
    benchmark = runBenchmark(p.getOVPath(), p, excelWriter, runOnDevice);    next(benchmark);    
    blobFilesProvider(p.getBlobFilesPath(), benchmark)        

if __name__ == "__main__":
    main()