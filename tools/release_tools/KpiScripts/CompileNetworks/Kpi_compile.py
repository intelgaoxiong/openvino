import os
import ipdb
import json
import glob
import subprocess
import colorama
import enum
from colorama import Fore
from subprocess import Popen, PIPE
from collections import namedtuple
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side



class Utilities:
    CONFIG = os.sys.argv[1]
    if not CONFIG.endswith(".json"):
        CONFIG = "./KpiScripts/CompileNetworks/compile_configs/compile_config.json"
    OUTPUT_BLOBS = os.getcwd() + "/CompiledModels"
    ERROR_STATUS = "FAIL"
    TIMEOUT = 420 

    class ExcelHeader(enum.Enum):
        BLOB_NAME = 1
        PACKAGE = 2
        FRAMEWORK = 3
        COMPILER = 4
        RESULT = 5
        MESSAGE = 6

    
    @classmethod
    def extractCompileAttrs(cls, jsonDict) -> str :     
        attrs="";
        for element in jsonDict.items():
            if element[0] == "attrs":
               attrs = "".join([" ".join([" -"+pair[0], pair[1]+" "]) for pair in element[1].items()])
               break;
        return attrs

    @classmethod
    def extractPaths(cls, jsonDict, ov_path) -> namedtuple :
        Path = namedtuple("Path", "ov_path ir_files_path")
        result = None

        for element in jsonDict.items():
            if element[0] == "paths":
                if ov_path is not None:
                    result = Path(ov_path, element[1]["ir_files_path"])                
                else:
                    result = Path(element[1]["ov_path"], element[1]["ir_files_path"])                
        return result

    @classmethod
    def extractCompilerType(cls, jsonDict) -> str :
        for element in jsonDict.items():
            if element[0] == "compiler":
                return element[1]["compilerType"]                

    @classmethod
    def pathConcat(cls, arg_basePath, arg_secondaryPath):
        fpath = (arg_basePath[-1] == '/')
        spath = (arg_secondaryPath[0] == '/')

        if fpath:
            if spath:
                return arg_basePath + arg_secondaryPath[1:];
            else:
                return arg_basePath + arg_secondaryPath;
        else:
            if spath:
                return arg_basePath + arg_secondaryPath;
            else:
                return arg_basePath + "/" + arg_secondaryPath;

    @classmethod
    def createErrFile(cls, irFile, message:str, arg_compilerType:str):        
        errFilePath = ".".join(irFile.split(".")[:-1]) + "_" + arg_compilerType + ".err"        
        with open(errFilePath, "w") as errFile:
            errFile.write(message)

    @classmethod
    def help(cls):
        print("This is a description of compile_config.json file !")
        print("attrs : every k,v pair will be forwarded to compile_tool")
        print("paths: path to private openvino and to POR models (current script is dependent of POR folder structure")
        print("compiler: set the compiler for compile_tool application, options : MLIR or MCM")

    def insertNewLines(arg_message:str, arg_colWidth:int):
        nrOfInsertedNewLines = 0
        divider = int(len(arg_message) / arg_colWidth) + 1
        offset = int(len(arg_message) / divider)        
        for pos in range(offset, len(arg_message), offset):             
            arg_message = arg_message[:pos] + "\n" + arg_message[pos:]
            nrOfInsertedNewLines += 1
        return (nrOfInsertedNewLines,arg_message)

# Path to POR models should be provided
def irFilesProvider(irFilesPath):
    # pathname = Utilities.pathConcat(irFilesPath,"/*/*/*/FP16-INT8/*.xml")
    print("xing test irFilesPath",irFilesPath)
    pathname = Utilities.pathConcat(irFilesPath, "/**/optimized/*.xml")   
    print("xing test pathname",pathname) 
    files = glob.glob(pathname, recursive=True)
    print("xing test files",files) 
    # ipdb.set_trace()
    for file in files:
        yield file

def excelWriter(arg_fileName : str):
    wb = Workbook()
    ws = wb.active
    entries = []      
    saved = False  
    styles = [{"alignment":Alignment(horizontal='center'), "font":Font(bold=True), "fill":PatternFill("solid", fgColor="D2D1D1"), "border":Border(top=Side(border_style="thin", color="000000"), left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))},
              {"alignment":Alignment(horizontal='center')},
              {"fill":PatternFill("solid", fgColor="D2D1D1"), "font.color.index":"FFFF0000"}]        
    BIGGER_THAN_TEXT = 1.2
    ERROR_MESSAGE_WIDTH = 125
    HEIGHT_OF_ROW = 21
    nextFreeRow = 2

    # header = ["BlobName", "Result", "Message"];
    header = {
                Utilities.ExcelHeader.BLOB_NAME : "ModelName",
                Utilities.ExcelHeader.PACKAGE : "Package",
                Utilities.ExcelHeader.FRAMEWORK : "FW",
                Utilities.ExcelHeader.COMPILER : "Compiler",
                Utilities.ExcelHeader.RESULT : "Result",
                Utilities.ExcelHeader.MESSAGE : "Message"
             }
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.BLOB_NAME.value)].width = 50
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.PACKAGE.value)].width = 15
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.FRAMEWORK.value)].width = 15
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.COMPILER.value)].width = 15
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.RESULT.value)].width = 15
    ws.column_dimensions[get_column_letter(Utilities.ExcelHeader.MESSAGE.value)].width = ERROR_MESSAGE_WIDTH    


    # Create excel header    
    for idx, e in enumerate(Utilities.ExcelHeader):
        # if int(ws.column_dimensions[get_column_letter(idx+1)].width * BIGGER_THAN_TEXT) < int(len(str(header[idx])) * BIGGER_THAN_TEXT):
        #     ws.column_dimensions[get_column_letter(idx+1)].width = int(len(header[idx]) * BIGGER_THAN_TEXT)        
        cell = ws.cell(column=idx+1, row=nextFreeRow, value=header[e]);
        for k, v in styles[0].items():
            setattr(cell, str(k), v);
    nextFreeRow = nextFreeRow + 1
    
    try:
        while True:
            entry = yield              
            ok = True;

            # entry = {} like header
            #Strip all new lines in message
            entry[Utilities.ExcelHeader.MESSAGE] = entry[Utilities.ExcelHeader.MESSAGE].replace("\n","")
            nrOfNewLines, entry[Utilities.ExcelHeader.MESSAGE] = Utilities.insertNewLines(entry[Utilities.ExcelHeader.MESSAGE], ERROR_MESSAGE_WIDTH)            
            ws.row_dimensions[nextFreeRow].height = HEIGHT_OF_ROW *nrOfNewLines

            for idx, value in enumerate(Utilities.ExcelHeader):
                cell = ws.cell(column=idx+1, row=nextFreeRow, value=entry[value]);
                cell.alignment = Alignment(vertical='center', horizontal='center')
                if value == Utilities.ExcelHeader.RESULT and entry[value] == Utilities.ERROR_STATUS:                    
                    cell.font = Font(color="FFFF0000")                    

            nextFreeRow = nextFreeRow + 1            
    finally:
        wb.save(filename = arg_fileName)
        


#Receives path to config file with compiler attributes, ov_path and ir_files_path
def compileRunner(configFilePath, ov_path):
        
    PATH_TO_COMPILER_TYPE = "/home/kmb/openvino_kmb_2021.4_extras/models/compile_for_kmb.conf"

    with open(configFilePath) as compileConfig:
        data = json.load(compileConfig)
        compileAttrs = Utilities.extractCompileAttrs(data)
        paths = Utilities.extractPaths(data, ov_path)   
        compilerType = Utilities.extractCompilerType(data)

    with open(PATH_TO_COMPILER_TYPE, "w") as compilerConfigType:
        compilerConfigType.write("VPUX_COMPILER_TYPE "+str(compilerType))

    irFiles = irFilesProvider(paths.ir_files_path)

    # Run compile_tool for every IR file and yield output
    for irFile in irFiles:               
        #Create excelEntry based on the template specified in ExcelHeader
        excelEntry = {e:None for e in Utilities.ExcelHeader}  

        #Drop the irFile xml extension
        modelName = ".".join([part for part in irFile.split('/')[-1].split(".")[:-1]])      
        
        #Update excelEntry with model name and compiler type (MLIR or MCM)
        #Update with framework and package extracted from file name
        excelEntry.update({Utilities.ExcelHeader.BLOB_NAME: modelName, Utilities.ExcelHeader.COMPILER: compilerType})
        excelEntry.update({
            Utilities.ExcelHeader.FRAMEWORK: irFile.split("/")[-3], 
            Utilities.ExcelHeader.PACKAGE: irFile.split("/")[-5]
            })

        try:                                   
            # Run compile_tool for irFile.xml
            process = subprocess.Popen('/bin/bash', stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)            
            #" -o " + Utilities.pathConcat(Utilities.OUTPUT_BLOBS, irFile.split('/')[-1].split('.')[0]+"_"+str(compilerType)+"_FP16_U8"+".blob" +                
            cmd = "cd "+paths.ov_path+"; source './bin/setupvars.sh' > /dev/null; ./deployment_tools/inference_engine/bin/compile_tool -il NCHW -ol NCHW -m "+ irFile + compileAttrs + " -c " + PATH_TO_COMPILER_TYPE + " -o " + ".".join(irFile.split(".")[:-1]) +"_"+ str(compilerType) + ".blob";
            print(Fore.GREEN + "[INFO] Compile " + str(irFile))
            output, err = process.communicate(cmd.encode(), timeout=Utilities.TIMEOUT);
            
            # Check the output
            if err != None and "Done. LoadNetwork time elapsed".encode() in output:
                #Compilation successful   
                            
                # entry = ExcelEntry(irFile.split('/')[-1].split(".")[0], "OK", "Compilation successful")                
                excelEntry.update({Utilities.ExcelHeader.RESULT: "OK", Utilities.ExcelHeader.MESSAGE:"Compilation successful"})
            else:
                #Compilation failed !                    
                Utilities.createErrFile(irFile, err.decode(), compilerType)                
                excelEntry.update({Utilities.ExcelHeader.RESULT: Utilities.ERROR_STATUS, Utilities.ExcelHeader.MESSAGE:err.decode()})
                print(Fore.RED+"       Compilation error !")                
            yield excelEntry
        except subprocess.TimeoutExpired:
            Utilities.createErrFile(irFile, "Timeout reached, probably infinite loop in compilation", compilerType)
            # entry = ExcelEntry(irFile.split('/')[-1].split(".")[0], Utilities.ERROR_STATUS, "Timeout reached, probably infinite loop in compilation")
            excelEntry.update({Utilities.ExcelHeader.RESULT: Utilities.ERROR_STATUS, Utilities.ExcelHeader.MESSAGE:"Timeout reached, probably infinite loop in compilation"})
            print(Fore.RED+"\tTimeout reached, probably infinite loop in compilation")
            yield excelEntry



def main(ov_path=None):
    compileOutput = compileRunner(Utilities.CONFIG, ov_path)  
    tempData = None    
    print(Utilities.CONFIG)
    with open(Utilities.CONFIG) as temp:
        tempData = json.load(temp)

    writeToExcel = excelWriter("./CompilingResults_"+Utilities.extractCompilerType(tempData)+".xlsx")
    # Initialize the excel writer
    writeToExcel.send(None)    
    for excelEntry in compileOutput:
        writeToExcel.send(excelEntry)
        # print(excelEntry)
    writeToExcel.close()

if __name__ == "__main__":
    main()
